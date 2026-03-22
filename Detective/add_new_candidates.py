"""Add new candidate entries to the xlsx and run the scraper on them."""

import openpyxl
import os

XLSX_PATH = os.path.join(os.path.dirname(__file__), "murder_mystery_candidates_v2.xlsx")

NEW_ENTRIES = [
    # === NOVELS ===

    # Cormoran Strike - Robert Galbraith
    ("The Cuckoo's Calling", "Robert Galbraith", 2013, "Novel", "Private Detective", "Explicit", "H", "Cormoran Strike", 1, ""),
    ("The Silkworm", "Robert Galbraith", 2014, "Novel", "Private Detective", "Explicit", "H", "Cormoran Strike", 2, ""),
    ("Career of Evil", "Robert Galbraith", 2015, "Novel", "Private Detective", "Explicit", "H", "Cormoran Strike", 3, ""),
    ("Lethal White", "Robert Galbraith", 2018, "Novel", "Private Detective", "Explicit", "H", "Cormoran Strike", 4, ""),
    ("Troubled Blood", "Robert Galbraith", 2020, "Novel", "Private Detective", "Explicit", "H", "Cormoran Strike", 5, ""),

    # Hardy Boys - Franklin W. Dixon
    ("The Tower Treasure", "Franklin W. Dixon", 1927, "Novel", "Juvenile Mystery", "Explicit", "M", "Hardy Boys", 1, ""),
    ("The House on the Cliff", "Franklin W. Dixon", 1927, "Novel", "Juvenile Mystery", "Explicit", "M", "Hardy Boys", 2, ""),
    ("The Secret of the Old Mill", "Franklin W. Dixon", 1927, "Novel", "Juvenile Mystery", "Explicit", "M", "Hardy Boys", 3, ""),
    ("The Missing Chums", "Franklin W. Dixon", 1928, "Novel", "Juvenile Mystery", "Explicit", "M", "Hardy Boys", 4, ""),
    ("Hunting for Hidden Gold", "Franklin W. Dixon", 1928, "Novel", "Juvenile Mystery", "Explicit", "M", "Hardy Boys", 5, ""),

    # Nancy Drew - Carolyn Keene
    ("The Secret of the Old Clock", "Carolyn Keene", 1930, "Novel", "Juvenile Mystery", "Explicit", "M", "Nancy Drew", 1, ""),
    ("The Hidden Staircase", "Carolyn Keene", 1930, "Novel", "Juvenile Mystery", "Explicit", "M", "Nancy Drew", 2, ""),
    ("The Bungalow Mystery", "Carolyn Keene", 1930, "Novel", "Juvenile Mystery", "Explicit", "M", "Nancy Drew", 3, ""),
    ("The Mystery at Lilac Inn", "Carolyn Keene", 1930, "Novel", "Juvenile Mystery", "Explicit", "M", "Nancy Drew", 4, ""),
    ("The Secret of Shadow Ranch", "Carolyn Keene", 1931, "Novel", "Juvenile Mystery", "Explicit", "M", "Nancy Drew", 5, ""),

    # Trixie Belden
    ("The Secret of the Mansion", "Julie Campbell", 1948, "Novel", "Juvenile Mystery", "Explicit", "M", "Trixie Belden", 1, ""),
    ("The Red Trailer Mystery", "Julie Campbell", 1950, "Novel", "Juvenile Mystery", "Explicit", "M", "Trixie Belden", 2, ""),
    ("The Gatehouse Mystery", "Julie Campbell", 1951, "Novel", "Juvenile Mystery", "Explicit", "M", "Trixie Belden", 3, ""),

    # Dana Girls
    ("By the Light of the Study Lamp", "Carolyn Keene", 1934, "Novel", "Juvenile Mystery", "Explicit", "M", "Dana Girls", 1, ""),
    ("The Secret at Lone Tree Cottage", "Carolyn Keene", 1934, "Novel", "Juvenile Mystery", "Explicit", "M", "Dana Girls", 2, ""),
    ("In the Shadow of the Tower", "Carolyn Keene", 1934, "Novel", "Juvenile Mystery", "Explicit", "M", "Dana Girls", 3, ""),

    # Penny Parker
    ("Tale of the Witch Doll", "Mildred Wirt Benson", 1939, "Novel", "Juvenile Mystery", "Explicit", "M", "Penny Parker", 1, ""),
    ("The Vanishing Houseboat", "Mildred Wirt Benson", 1939, "Novel", "Juvenile Mystery", "Explicit", "M", "Penny Parker", 2, ""),
    ("Danger at the Drawbridge", "Mildred Wirt Benson", 1940, "Novel", "Juvenile Mystery", "Explicit", "M", "Penny Parker", 3, ""),

    # Great Shelby Holmes
    ("The Great Shelby Holmes", "Elizabeth Eulberg", 2017, "Novel", "Juvenile Mystery", "Explicit", "M", "Great Shelby Holmes", 1, ""),
    ("The Great Shelby Holmes Meets Her Match", "Elizabeth Eulberg", 2018, "Novel", "Juvenile Mystery", "Explicit", "M", "Great Shelby Holmes", 2, ""),

    # Three Investigators
    ("The Secret of Terror Castle", "Robert Arthur", 1964, "Novel", "Juvenile Mystery", "Explicit", "M", "Three Investigators", 1, ""),
    ("The Mystery of the Stuttering Parrot", "Robert Arthur", 1964, "Novel", "Juvenile Mystery", "Explicit", "M", "Three Investigators", 2, ""),
    ("The Mystery of the Whispering Mummy", "Robert Arthur", 1965, "Novel", "Juvenile Mystery", "Explicit", "M", "Three Investigators", 3, ""),

    # Charlie Thorne
    ("Charlie Thorne and the Last Equation", "Stuart Gibbs", 2019, "Novel", "Juvenile Mystery", "Explicit", "M", "Charlie Thorne", 1, ""),
    ("Charlie Thorne and the Lost City", "Stuart Gibbs", 2021, "Novel", "Juvenile Mystery", "Explicit", "M", "Charlie Thorne", 2, ""),

    # Framed!
    ("Framed!", "James Ponti", 2016, "Novel", "Juvenile Mystery", "Explicit", "M", "City Spies / Framed", 1, ""),
    ("Vanished!", "James Ponti", 2017, "Novel", "Juvenile Mystery", "Explicit", "M", "City Spies / Framed", 2, ""),

    # 39 Clues
    ("The Maze of Bones", "Rick Riordan", 2008, "Novel", "Juvenile Mystery", "Explicit", "M", "The 39 Clues", 1, ""),
    ("One False Note", "Gordon Korman", 2008, "Novel", "Juvenile Mystery", "Explicit", "M", "The 39 Clues", 2, ""),
    ("The Sword Thief", "Peter Lerangis", 2009, "Novel", "Juvenile Mystery", "Explicit", "M", "The 39 Clues", 3, ""),

    # Peculiar Crimes Unit - Christopher Fowler
    ("Full Dark House", "Christopher Fowler", 2003, "Novel", "Procedural", "Explicit", "H", "Bryant & May", 1, ""),
    ("The Water Room", "Christopher Fowler", 2004, "Novel", "Procedural", "Explicit", "H", "Bryant & May", 2, ""),
    ("Seventy-Seven Clocks", "Christopher Fowler", 2005, "Novel", "Procedural", "Explicit", "H", "Bryant & May", 3, ""),
    ("Ten Second Staircase", "Christopher Fowler", 2006, "Novel", "Procedural", "Explicit", "H", "Bryant & May", 4, ""),

    # Flavia de Luce - Alan Bradley
    ("The Sweetness at the Bottom of the Pie", "Alan Bradley", 2009, "Novel", "Cozy Mystery", "Explicit", "H", "Flavia de Luce", 1, ""),
    ("The Weed That Strings the Hangman's Bag", "Alan Bradley", 2010, "Novel", "Cozy Mystery", "Explicit", "H", "Flavia de Luce", 2, ""),
    ("A Red Herring Without Mustard", "Alan Bradley", 2011, "Novel", "Cozy Mystery", "Explicit", "H", "Flavia de Luce", 3, ""),
    ("I Am Half-Sick of Shadows", "Alan Bradley", 2011, "Novel", "Cozy Mystery", "Explicit", "H", "Flavia de Luce", 4, ""),

    # Armand Gamache - Louise Penny
    ("Still Life", "Louise Penny", 2005, "Novel", "Village Mystery", "Explicit", "H", "Chief Inspector Gamache", 1, ""),
    ("A Fatal Grace", "Louise Penny", 2007, "Novel", "Village Mystery", "Explicit", "H", "Chief Inspector Gamache", 2, ""),
    ("The Cruelest Month", "Louise Penny", 2008, "Novel", "Village Mystery", "Explicit", "H", "Chief Inspector Gamache", 3, ""),
    ("A Rule Against Murder", "Louise Penny", 2009, "Novel", "Village Mystery", "Explicit", "H", "Chief Inspector Gamache", 4, ""),
    ("The Brutal Telling", "Louise Penny", 2009, "Novel", "Village Mystery", "Explicit", "H", "Chief Inspector Gamache", 5, ""),

    # Nero Wolfe - Rex Stout
    ("Fer-de-Lance", "Rex Stout", 1934, "Novel", "Classic Whodunit", "Explicit", "H", "Nero Wolfe", 1, ""),
    ("The League of Frightened Men", "Rex Stout", 1935, "Novel", "Classic Whodunit", "Explicit", "H", "Nero Wolfe", 2, ""),
    ("Too Many Cooks", "Rex Stout", 1938, "Novel", "Classic Whodunit", "Explicit", "H", "Nero Wolfe", 3, ""),
    ("Some Buried Caesar", "Rex Stout", 1939, "Novel", "Classic Whodunit", "Explicit", "H", "Nero Wolfe", 4, ""),

    # Alistair MacLean
    ("The Guns of Navarone", "Alistair MacLean", 1957, "Novel", "Thriller", "Explicit", "H", None, None, ""),
    ("Ice Station Zebra", "Alistair MacLean", 1963, "Novel", "Thriller", "Explicit", "H", None, None, ""),
    ("Where Eagles Dare", "Alistair MacLean", 1967, "Novel", "Thriller", "Explicit", "H", None, None, ""),
    ("Fear Is the Key", "Alistair MacLean", 1961, "Novel", "Thriller", "Explicit", "H", None, None, ""),
    ("Night Without End", "Alistair MacLean", 1959, "Novel", "Thriller", "Explicit", "H", None, None, ""),

    # Tony Hillerman
    ("The Blessing Way", "Tony Hillerman", 1970, "Novel", "Mystery", "Explicit", "H", "Navajo Mysteries", 1, ""),
    ("Dance Hall of the Dead", "Tony Hillerman", 1973, "Novel", "Mystery", "Explicit", "H", "Navajo Mysteries", 2, ""),
    ("Listening Woman", "Tony Hillerman", 1978, "Novel", "Mystery", "Explicit", "H", "Navajo Mysteries", 3, ""),
    ("Skinwalkers", "Tony Hillerman", 1986, "Novel", "Mystery", "Explicit", "H", "Navajo Mysteries", 4, ""),
    ("A Thief of Time", "Tony Hillerman", 1988, "Novel", "Mystery", "Explicit", "H", "Navajo Mysteries", 5, ""),

    # Travis McGee - John D. MacDonald
    ("The Deep Blue Good-by", "John D. MacDonald", 1964, "Novel", "Hardboiled", "Explicit", "H", "Travis McGee", 1, ""),
    ("Nightmare in Pink", "John D. MacDonald", 1964, "Novel", "Hardboiled", "Explicit", "H", "Travis McGee", 2, ""),
    ("A Purple Place for Dying", "John D. MacDonald", 1964, "Novel", "Hardboiled", "Explicit", "H", "Travis McGee", 3, ""),
    ("The Lonely Silver Rain", "John D. MacDonald", 1985, "Novel", "Hardboiled", "Explicit", "H", "Travis McGee", 4, ""),

    # Dick Francis
    ("Dead Cert", "Dick Francis", 1962, "Novel", "Mystery", "Explicit", "H", None, None, "Horse racing mystery"),
    ("Nerve", "Dick Francis", 1964, "Novel", "Mystery", "Explicit", "H", None, None, "Horse racing mystery"),
    ("Odds Against", "Dick Francis", 1965, "Novel", "Mystery", "Explicit", "H", None, None, "Horse racing mystery; first Sid Halley"),
    ("Whip Hand", "Dick Francis", 1979, "Novel", "Mystery", "Explicit", "H", None, None, "Horse racing mystery"),
    ("Proof", "Dick Francis", 1984, "Novel", "Mystery", "Explicit", "H", None, None, "Horse racing mystery"),

    # A Good Girl's Guide to Murder (novel)
    ("A Good Girl's Guide to Murder", "Holly Jackson", 2019, "Novel", "Young Adult Mystery", "Explicit", "H", None, None, ""),

    # === TV SHOWS ===

    # Veronica Mars
    ("Veronica Mars (season 1)", "Rob Thomas", 2004, "TV Episode", "Teen Mystery", "Explicit", "H", "Veronica Mars", 1, ""),
    ("Veronica Mars (season 2)", "Rob Thomas", 2005, "TV Episode", "Teen Mystery", "Explicit", "H", "Veronica Mars", 2, ""),
    ("Veronica Mars (season 3)", "Rob Thomas", 2006, "TV Episode", "Teen Mystery", "Explicit", "H", "Veronica Mars", 3, ""),
    ("Veronica Mars (season 4)", "Rob Thomas", 2019, "TV Episode", "Teen Mystery", "Explicit", "H", "Veronica Mars", 4, ""),

    # Veronica Mars film
    ("Veronica Mars", "Rob Thomas", 2014, "Film", "Teen Mystery", "Explicit", "H", "Veronica Mars", None, ""),

    # Elementary (selected seasons)
    ("Elementary (season 1)", "Robert Doherty", 2012, "TV Episode", "Modern Whodunit", "Explicit", "H", "Elementary", 1, ""),
    ("Elementary (season 2)", "Robert Doherty", 2013, "TV Episode", "Modern Whodunit", "Explicit", "H", "Elementary", 2, ""),
    ("Elementary (season 3)", "Robert Doherty", 2014, "TV Episode", "Modern Whodunit", "Explicit", "H", "Elementary", 3, ""),

    # A Good Girl's Guide to Murder (TV)
    ("A Good Girl's Guide to Murder (TV series)", "Holly Jackson", 2024, "TV Episode", "Young Adult Mystery", "Explicit", "H", None, None, ""),

    # Stumptown
    ("Stumptown (TV series)", "Jason Richman", 2019, "TV Episode", "Hardboiled", "Explicit", "H", None, None, ""),

    # Psych (selected seasons)
    ("Psych (season 1)", "Steve Franks", 2006, "TV Episode", "Comedy Mystery", "Explicit", "M", "Psych", 1, ""),
    ("Psych (season 2)", "Steve Franks", 2007, "TV Episode", "Comedy Mystery", "Explicit", "M", "Psych", 2, ""),
    ("Psych (season 3)", "Steve Franks", 2008, "TV Episode", "Comedy Mystery", "Explicit", "M", "Psych", 3, ""),

    # Bones (selected seasons)
    ("Bones (season 1)", "Hart Hanson", 2005, "TV Episode", "Forensic Procedural", "Explicit", "M", "Bones", 1, ""),
    ("Bones (season 2)", "Hart Hanson", 2006, "TV Episode", "Forensic Procedural", "Explicit", "M", "Bones", 2, ""),
    ("Bones (season 3)", "Hart Hanson", 2007, "TV Episode", "Forensic Procedural", "Explicit", "M", "Bones", 3, ""),

    # Nancy Drew (CW TV)
    ("Nancy Drew (TV series) season 1", "Noga Landau", 2019, "TV Episode", "Supernatural Mystery", "Explicit", "H", "Nancy Drew TV", 1, ""),
    ("Nancy Drew (TV series) season 2", "Noga Landau", 2021, "TV Episode", "Supernatural Mystery", "Explicit", "H", "Nancy Drew TV", 2, ""),

    # Monk (selected seasons)
    ("Monk (season 1)", "Andy Breckman", 2002, "TV Episode", "Comedy Mystery", "Explicit", "M", "Monk", 1, ""),
    ("Monk (season 2)", "Andy Breckman", 2003, "TV Episode", "Comedy Mystery", "Explicit", "M", "Monk", 2, ""),
    ("Monk (season 3)", "Andy Breckman", 2004, "TV Episode", "Comedy Mystery", "Explicit", "M", "Monk", 3, ""),

    # CSI (selected seasons)
    ("CSI: Crime Scene Investigation (season 1)", "Anthony E. Zuiker", 2000, "TV Episode", "Forensic Procedural", "Explicit", "M", "CSI", 1, ""),
    ("CSI: Crime Scene Investigation (season 2)", "Anthony E. Zuiker", 2001, "TV Episode", "Forensic Procedural", "Explicit", "M", "CSI", 2, ""),

    # Prodigal Son
    ("Prodigal Son (season 1)", "Chris Fedak", 2019, "TV Episode", "Psychological Thriller", "Explicit", "H", "Prodigal Son", 1, ""),
    ("Prodigal Son (season 2)", "Chris Fedak", 2021, "TV Episode", "Psychological Thriller", "Explicit", "H", "Prodigal Son", 2, ""),

    # Unbelievable
    ("Unbelievable (miniseries)", "Susannah Grant", 2019, "TV Episode", "True Crime", "Explicit", "H", None, None, "Netflix miniseries"),

    # Murdoch Mysteries (selected seasons)
    ("Murdoch Mysteries (season 1)", "Cal Coons", 2008, "TV Episode", "Historical Mystery", "Explicit", "H", "Murdoch Mysteries", 1, ""),
    ("Murdoch Mysteries (season 2)", "Cal Coons", 2009, "TV Episode", "Historical Mystery", "Explicit", "H", "Murdoch Mysteries", 2, ""),
    ("Murdoch Mysteries (season 3)", "Cal Coons", 2010, "TV Episode", "Historical Mystery", "Explicit", "H", "Murdoch Mysteries", 3, ""),

    # Young Sherlock Holmes
    ("Young Sherlock Holmes", "Barry Levinson", 1985, "Film", "Classic Whodunit", "Explicit", "H", None, None, ""),

    # The Marlow Murder Club
    ("The Marlow Murder Club (TV series)", "Robert Thorogood", 2024, "TV Episode", "Cozy Mystery", "Explicit", "H", None, None, ""),

    # The Fall
    ("The Fall (series 1)", "Allan Cubitt", 2013, "TV Episode", "Psychological Thriller", "Explicit", "H", "The Fall", 1, "BBC"),
    ("The Fall (series 2)", "Allan Cubitt", 2014, "TV Episode", "Psychological Thriller", "Explicit", "H", "The Fall", 2, "BBC"),
    ("The Fall (series 3)", "Allan Cubitt", 2016, "TV Episode", "Psychological Thriller", "Explicit", "H", "The Fall", 3, "BBC"),

    # Inspector Lynley Mysteries
    ("The Inspector Lynley Mysteries", "Elizabeth George", 2001, "TV Episode", "British Procedural", "Explicit", "H", None, None, "BBC"),

    # Shetland (selected series)
    ("Shetland (series 1)", "Ann Cleeves", 2013, "TV Episode", "British Procedural", "Explicit", "H", "Shetland", 1, ""),
    ("Shetland (series 2)", "Ann Cleeves", 2014, "TV Episode", "British Procedural", "Explicit", "H", "Shetland", 2, ""),
    ("Shetland (series 3)", "Ann Cleeves", 2016, "TV Episode", "British Procedural", "Explicit", "H", "Shetland", 3, ""),
    ("Shetland (series 4)", "Ann Cleeves", 2018, "TV Episode", "British Procedural", "Explicit", "H", "Shetland", 4, ""),
    ("Shetland (series 5)", "Ann Cleeves", 2019, "TV Episode", "British Procedural", "Explicit", "H", "Shetland", 5, ""),

    # Midsomer Murders (additional series-level entries)
    ("Midsomer Murders (series 6)", "Various", 2003, "TV Episode", "Cozy Mystery", "Explicit", "H", "Midsomer Murders", 6, ""),
    ("Midsomer Murders (series 7)", "Various", 2004, "TV Episode", "Cozy Mystery", "Explicit", "H", "Midsomer Murders", 7, ""),
    ("Midsomer Murders (series 8)", "Various", 2005, "TV Episode", "Cozy Mystery", "Explicit", "H", "Midsomer Murders", 8, ""),
    ("Midsomer Murders (series 9)", "Various", 2006, "TV Episode", "Cozy Mystery", "Explicit", "H", "Midsomer Murders", 9, ""),
    ("Midsomer Murders (series 10)", "Various", 2007, "TV Episode", "Cozy Mystery", "Explicit", "H", "Midsomer Murders", 10, ""),
]

HEADERS = [
    "Title",
    "Author / Director",
    "Year",
    "Medium",
    "Subgenre",
    "Villain Reveal",
    "Synopsis Quality",
    "Series Name",
    "Series Entry #",
    "Notes",
]


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Candidate List"]

    # Get existing titles to avoid duplicates
    existing = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0]:
            existing.add(row[0].strip().lower())

    added = 0
    skipped = 0
    for entry in NEW_ENTRIES:
        title = entry[0].strip()
        if title.lower() in existing:
            print(f"  SKIP (duplicate): {title}")
            skipped += 1
            continue

        ws.append(list(entry))
        existing.add(title.lower())
        added += 1

    wb.save(XLSX_PATH)
    print(f"\nAdded {added} new entries, skipped {skipped} duplicates.")
    print(f"Total rows now: {ws.max_row - 1}")


if __name__ == "__main__":
    main()
